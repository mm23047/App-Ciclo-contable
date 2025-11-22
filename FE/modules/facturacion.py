"""
Módulo Streamlit para Facturación Digital.
Sistema de emisión de facturas electrónicas con integración contable.
"""
import streamlit as st
import requests
import pandas as pd
from datetime import datetime, date
from typing import Dict, Any, List
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import json
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl.styles import Font, PatternFill, Alignment

def render_page(backend_url: str):
    """Renderizar página de facturación"""
    
    st.header("🧾 Facturación Digital")
    st.markdown("Sistema de emisión de facturas electrónicas con integración contable automática")
    
    # Tabs para organizar funcionalidades
    tab1, tab2, tab3, tab4 = st.tabs(["📝 Nueva Factura", "📋 Gestión de Facturas", "📊 Reportes de Ventas", "⚙️ Configuración"])
    
    with tab1:
        crear_nueva_factura(backend_url)
    
    with tab2:
        gestion_facturas(backend_url)
    
    with tab3:
        reportes_ventas(backend_url)
    
    with tab4:
        configuracion_facturacion(backend_url)

def crear_nueva_factura(backend_url: str):
    """Crear nueva factura"""
    
    st.subheader("📝 Crear Nueva Factura")
    
    # Cargar configuración fiscal
    try:
        response_config = requests.get(f"{backend_url}/api/facturacion/configuracion")
        if response_config.status_code == 200:
            config = response_config.json()
            
            # Mostrar información fiscal activa
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("IVA", f"{float(config['iva_porcentaje']):.1f}%")
            with col2:
                st.metric("Ret. Fuente", f"{float(config['retefuente_porcentaje']):.2f}%")
            with col3:
                st.metric("ReteICA", f"{float(config['reteica_porcentaje']):.3f}%")
            with col4:
                st.metric("Próximo #", f"{config['prefijo_factura']}-{str(config['numero_actual']).zfill(4)}")
            
            st.divider()
        else:
            config = None
    except Exception as e:
        config = None
        st.warning(f"⚠️ Error al cargar configuración fiscal: {e}")
    
    # Obtener clientes disponibles
    try:
        response_clientes = requests.get(f"{backend_url}/api/facturacion/clientes")
        clientes = response_clientes.json() if response_clientes.status_code == 200 else []
    except:
        clientes = []
    
    # Obtener productos disponibles
    try:
        response_productos = requests.get(f"{backend_url}/api/facturacion/productos")
        productos = response_productos.json() if response_productos.status_code == 200 else []
    except:
        productos = []
    
    if not clientes:
        st.warning("⚠️ No hay clientes registrados. Ve a Configuración para agregar clientes.")
        return
    
    if not productos:
        st.warning("⚠️ No hay productos registrados. Ve a Configuración para agregar productos.")
        return
    
    # Información básica de la factura
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### 👤 Información del Cliente")
        
        # Selección de cliente
        opciones_clientes = [
            f"{c['codigo_cliente']} - {c['nombre']} ({c['tipo_cliente']})"
            for c in clientes if c.get('estado_cliente') == 'ACTIVO'
        ]
        
        cliente_seleccionado = st.selectbox("Cliente:", opciones_clientes, key="sel_cliente_factura")
        
        # Obtener datos del cliente seleccionado
        cliente_obj = None
        if cliente_seleccionado:
            codigo_cliente = cliente_seleccionado.split(" - ")[0]
            cliente_obj = next((c for c in clientes if c['codigo_cliente'] == codigo_cliente), None)
            
            if cliente_obj:
                st.text_input("NIT/CC:", value=cliente_obj.get('nit', ''), disabled=True)
                st.text_input("Email:", value=cliente_obj.get('email', ''), disabled=True)
                st.text_area("Dirección:", value=cliente_obj.get('direccion', ''), disabled=True, height=60)
    
    with col2:
        st.markdown("#### 🧾 Datos de la Factura")
        
        # Datos de la factura
        fecha_factura = st.date_input(
            "Fecha de factura:",
            value=datetime.now().date(),
            key="fecha_factura"
        )
        
        fecha_vencimiento = st.date_input(
            "Fecha de vencimiento:",
            value=datetime.now().date(),
            key="fecha_vencimiento"
        )
        
        tipo_factura = st.selectbox(
            "Tipo de factura:",
            ["Contado", "Crédito"],
            key="tipo_factura"
        )
        
        observaciones = st.text_area(
            "Observaciones:",
            height=80,
            help="Observaciones adicionales para la factura",
            key="obs_factura"
        )
    
    # Sección de productos
    st.markdown("### 🛍️ Productos/Servicios")
    
    # Gestión de productos de la factura
    if 'productos_factura' not in st.session_state:
        st.session_state.productos_factura = []
    
    # Formulario para agregar productos
    with st.expander("➕ Agregar Producto/Servicio", expanded=True):
        col1, col2, col3 = st.columns([4, 1.5, 1.5])
        
        with col1:
            opciones_productos = [
                f"{p['codigo_producto']} - {p['nombre']}"
                for p in productos if p.get('estado_producto') == 'ACTIVO'
            ]
            
            if opciones_productos:
                producto_sel = st.selectbox("Producto/Servicio:", opciones_productos, key="prod_factura")
                
                # Mostrar precio del producto seleccionado
                if producto_sel:
                    codigo_prod = producto_sel.split(" - ")[0]
                    prod_obj = next((p for p in productos if p['codigo_producto'] == codigo_prod), None)
                    if prod_obj:
                        precio_producto = float(prod_obj.get('precio_venta', 0))
                        st.info(f"💰 Precio unitario: **${precio_producto:,.2f}**")
            else:
                st.warning("No hay productos activos")
                producto_sel = None
        
        with col2:
            cantidad = st.number_input("Cantidad:", min_value=0.01, value=1.0, step=0.01, key="cant_factura")
        
        with col3:
            descuento = st.number_input("Descuento %:", min_value=0.0, max_value=100.0, value=0.0, step=0.1, key="desc_factura")
        
        # Botón de agregar
        col_btn1, col_btn2 = st.columns([3, 1])
        with col_btn2:
            if st.button("➕ Agregar Producto", key="add_prod_factura", use_container_width=True, type="primary"):
                if producto_sel and cantidad > 0:
                    codigo_prod = producto_sel.split(" - ")[0]
                    prod_obj = next((p for p in productos if p['codigo_producto'] == codigo_prod), None)
                    
                    if prod_obj:
                        precio_unitario = float(prod_obj.get('precio_venta', 0))
                        subtotal = cantidad * precio_unitario * (1 - descuento/100)
                        
                        nuevo_item = {
                            'id_producto': prod_obj['id_producto'],
                            'codigo_producto': codigo_prod,
                            'nombre_producto': prod_obj['nombre'],
                            'cantidad': cantidad,
                            'precio_unitario': precio_unitario,
                            'descuento_porcentaje': descuento,
                            'subtotal': subtotal
                        }
                        
                        st.session_state.productos_factura.append(nuevo_item)
                        st.rerun()
                else:
                    st.error("⚠️ Selecciona un producto y una cantidad válida")
    
    # Mostrar productos agregados
    if st.session_state.productos_factura:
        st.markdown("#### 📋 Productos en la Factura")
        
        # Obtener porcentaje de IVA de configuración
        iva_porcentaje = float(config.get('iva_porcentaje', 13.0)) / 100 if config else 0.13
        
        # Estilos CSS para la tabla de productos
        st.markdown("""
        <style>
        .productos-table {
            width: 100%;
            border-collapse: collapse;
            margin: 15px 0;
            font-size: 14px;
            background-color: #ffffff;
        }
        .productos-table thead {
            background-color: #ffffff;
        }
        .productos-table th {
            padding: 12px 8px;
            text-align: left;
            font-weight: 600;
            border-bottom: 2px solid #ddd;
            color: #000000;
            background-color: #ffffff;
        }
        .productos-table td {
            padding: 10px 8px;
            border-bottom: 1px solid #eee;
            color: #000000;
            background-color: #ffffff;
        }
        .productos-table tbody tr {
            background-color: #ffffff;
        }
        .productos-table tbody tr:hover {
            background-color: #f8f9fa;
        }
        .text-right {
            text-align: right;
        }
        .text-center {
            text-align: center;
        }
        .producto-codigo {
            color: #000000;
            font-size: 11px;
        }
        .producto-nombre {
            font-weight: 500;
            color: #000000;
        }
        .descuento-text {
            color: #000000;
        }
        .iva-text {
            color: #000000;
        }
        .total-bold {
            font-weight: 600;
            color: #000000;
        }
        </style>
        """, unsafe_allow_html=True)
        
        # Construir tabla HTML
        html_productos = """<table class="productos-table">
<thead>
<tr>
    <th>Producto</th>
    <th class="text-right">Cantidad</th>
    <th class="text-right">Precio Unit.</th>
    <th class="text-right">Desc.</th>
    <th class="text-right">Subtotal</th>
    <th class="text-right">IVA</th>
    <th class="text-right">Total Línea</th>
</tr>
</thead>
<tbody>"""
        
        # Agregar filas de productos
        for item in st.session_state.productos_factura:
            iva_linea = item['subtotal'] * iva_porcentaje
            total_linea = item['subtotal'] + iva_linea
            
            descuento_display = f"<span class='descuento-text'>{item['descuento_porcentaje']:.1f}%</span>" if item['descuento_porcentaje'] > 0 else "-"
            
            html_productos += f"""
<tr>
    <td>
        <div class="producto-codigo">{item['codigo_producto']}</div>
        <div class="producto-nombre">{item['nombre_producto']}</div>
    </td>
    <td class="text-right">{item['cantidad']:,.2f}</td>
    <td class="text-right">${item['precio_unitario']:,.2f}</td>
    <td class="text-right">{descuento_display}</td>
    <td class="text-right">${item['subtotal']:,.2f}</td>
    <td class="text-right iva-text">${iva_linea:,.2f}</td>
    <td class="text-right total-bold">${total_linea:,.2f}</td>
</tr>"""
        
        html_productos += """
</tbody>
</table>"""
        
        st.markdown(html_productos, unsafe_allow_html=True)
        
        # Botones para eliminar productos
        st.markdown("")
        cols_delete = st.columns(len(st.session_state.productos_factura) if len(st.session_state.productos_factura) <= 5 else 5)
        for i, item in enumerate(st.session_state.productos_factura):
            col_idx = i if i < 5 else i % 5
            with cols_delete[col_idx]:
                if st.button(f"🗑️ {item['nombre_producto'][:15]}...", key=f"del_prod_{i}", help=f"Eliminar {item['nombre_producto']}"):
                    st.session_state.productos_factura.pop(i)
                    st.rerun()
        
        st.markdown("---")
    
    # Totales de la factura usando configuración fiscal
    if st.session_state.productos_factura:
        subtotal_total = sum(item['subtotal'] for item in st.session_state.productos_factura)
        
        # Usar porcentajes de la configuración
        if config:
            iva_porcentaje = float(config.get('iva_porcentaje', 13.0)) / 100
            retefuente_porcentaje = float(config.get('retefuente_porcentaje', 0.0)) / 100
            reteica_porcentaje = float(config.get('reteica_porcentaje', 0.0)) / 100
        else:
            iva_porcentaje = 0.13
            retefuente_porcentaje = 0.0
            reteica_porcentaje = 0.0
        
        iva_total = subtotal_total * iva_porcentaje
        retencion_fuente = subtotal_total * retefuente_porcentaje
        reteica = subtotal_total * reteica_porcentaje
        total_factura = subtotal_total + iva_total - retencion_fuente - reteica
        
        st.markdown("### 💰 Resumen de Totales")
        
        # Diseño en dos columnas: Detalle y Total
        col_detalle, col_total = st.columns([3, 1])
        
        with col_detalle:
            # Tabla de cálculos con estilos consistentes
            st.markdown("""
<style>
.totals-table {
    width: 100%;
    border-collapse: collapse;
    margin: 10px 0;
    font-size: 14px;
    background-color: #ffffff;
}
.totals-table tr {
    border-bottom: 1px solid #eee;
    background-color: #ffffff;
}
.totals-table tr:hover {
    background-color: #f8f9fa;
}
.totals-table td {
    padding: 10px 15px;
    color: #000000;
    background-color: #ffffff;
}
.totals-table .label {
    text-align: left;
    font-weight: 500;
    font-size: 14px;
    color: #000000;
}
.totals-table .value {
    text-align: right;
    font-family: 'Courier New', monospace;
    font-size: 14px;
    color: #000000;
}
.totals-table .total-row {
    background-color: #ffffff;
    font-weight: bold;
    font-size: 16px;
}
.totals-table .total-row td {
    padding: 15px;
    border-top: 2px solid #ddd;
    color: #000000;
    background-color: #ffffff;
}
.tax-positive {
    color: #000000 !important;
}
.tax-negative {
    color: #ff6b6b !important;
}
</style>
""", unsafe_allow_html=True)
            
            # Construir la tabla de totales
            html_table = f"""<table class="totals-table">
<tr>
    <td class="label">📦 Subtotal ({len(st.session_state.productos_factura)} items)</td>
    <td class="value">${subtotal_total:,.2f}</td>
</tr>
<tr>
    <td class="label tax-positive">➕ IVA ({iva_porcentaje*100:.1f}%)</td>
    <td class="value tax-positive">+${iva_total:,.2f}</td>
</tr>
"""
            
            # Agregar retenciones si aplican
            if retencion_fuente > 0:
                html_table += f"""<tr>
    <td class="label tax-negative">➖ Retención en la Fuente ({retefuente_porcentaje*100:.2f}%)</td>
    <td class="value tax-negative">-${retencion_fuente:,.2f}</td>
</tr>
"""
            
            if reteica > 0:
                html_table += f"""<tr>
    <td class="label tax-negative">➖ ReteICA ({reteica_porcentaje*100:.3f}%)</td>
    <td class="value tax-negative">-${reteica:,.2f}</td>
</tr>
"""
            
            html_table += f"""<tr class="total-row">
    <td class="label">💵 TOTAL A PAGAR</td>
    <td class="value">${total_factura:,.2f}</td>
</tr>
</table>"""
            
            st.markdown(html_table, unsafe_allow_html=True)
        
        with col_total:
            # Total destacado
            st.markdown("""
            <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        padding: 30px; 
                        border-radius: 10px; 
                        text-align: center; 
                        color: white;
                        box-shadow: 0 4px 6px rgba(0,0,0,0.1);'>
                <div style='font-size: 0.9em; opacity: 0.9; margin-bottom: 5px;'>Total a Pagar</div>
                <div style='font-size: 2em; font-weight: bold;'>$""" + f"{total_factura:,.2f}" + """</div>
            </div>
            """, unsafe_allow_html=True)
    
    # Botón para crear factura
    st.markdown("---")
    col1, col2 = st.columns([1, 1])
    
    with col1:
        if st.button("🧾 Crear Factura", width="stretch", type="primary", key="btn_crear_factura"):
            if not st.session_state.productos_factura:
                st.error("❌ Debe agregar al menos un producto/servicio")
            elif not cliente_seleccionado:
                st.error("❌ Debe seleccionar un cliente")
            else:
                crear_factura_backend(
                    backend_url,
                    cliente_obj,
                    fecha_factura,
                    fecha_vencimiento,
                    tipo_factura,
                    observaciones,
                    st.session_state.productos_factura
                )
    
    with col2:
        if st.button("🔄 Limpiar Factura", width="stretch", key="btn_limpiar_factura"):
            st.session_state.productos_factura = []
            st.rerun()

def crear_factura_backend(
    backend_url: str,
    cliente: Dict,
    fecha_factura: date,
    fecha_vencimiento: date,
    tipo_factura: str,
    observaciones: str,
    productos: List[Dict]
):
    """Enviar factura al backend"""
    
    try:
        # Preparar detalles (líneas de la factura)
        detalles = [
            {
                "numero_linea": idx + 1,
                "id_producto": item['id_producto'],
                "descripcion_personalizada": None,
                "cantidad": float(item['cantidad']),
                "precio_unitario": float(item['precio_unitario']),
                "descuento_linea": float(item['descuento_porcentaje'])
            }
            for idx, item in enumerate(productos)
        ]
        
        # Preparar datos completos de la factura (incluyendo detalles)
        datos_factura = {
            "serie_factura": "A",
            "fecha_emision": fecha_factura.isoformat(),
            "fecha_vencimiento": fecha_vencimiento.isoformat(),
            "id_cliente": cliente['id_cliente'],
            "metodo_pago": tipo_factura,
            "observaciones": observaciones if observaciones else None,
            "usuario_creacion": "SISTEMA",
            "detalles": detalles
        }
        
        with st.spinner("Creando factura..."):
            response = requests.post(
                f"{backend_url}/api/facturacion/facturas",
                json=datos_factura
            )
        
        if response.status_code in [200, 201]:
            factura_creada = response.json()
            st.success(f"✅ Factura creada exitosamente! Número: {factura_creada.get('numero_factura', 'N/A')}")
            st.session_state.productos_factura = []  # Limpiar productos
            
            # Mostrar resumen de la factura
            with st.expander("📄 Resumen de la Factura Creada", expanded=True):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write(f"**Número:** {factura_creada.get('numero_factura', 'N/A')}")
                    st.write(f"**Cliente:** {cliente['nombre']}")
                    st.write(f"**Fecha:** {fecha_factura}")
                    st.write(f"**Tipo:** {tipo_factura}")
                
                with col2:
                    subtotal = float(factura_creada.get('subtotal', 0))
                    iva = float(factura_creada.get('impuesto_iva', 0))
                    retencion_fuente = float(factura_creada.get('retencion_fuente', 0))
                    reteica = float(factura_creada.get('reteica', 0))
                    total = float(factura_creada.get('total', 0))
                    
                    st.write(f"**Subtotal:** ${subtotal:,.2f}")
                    st.write(f"**IVA:** +${iva:,.2f}")
                    if retencion_fuente > 0:
                        st.write(f"**Retención Fuente:** -${retencion_fuente:,.2f}")
                    if reteica > 0:
                        st.write(f"**ReteICA:** -${reteica:,.2f}")
                    st.write(f"**Total:** ${total:,.2f}")
                    st.write(f"**Estado:** {factura_creada.get('estado_factura', 'EMITIDA')}")
        else:
            try:
                error_detail = response.json().get('detail', 'Error desconocido')
            except:
                error_detail = f"Error HTTP {response.status_code}"
            st.error(f"❌ Error al crear factura: {error_detail}")
            
    except requests.exceptions.RequestException as e:
        st.error(f"❌ Error de conexión: {e}")
    except Exception as e:
        st.error(f"❌ Error inesperado: {e}")

def gestion_facturas(backend_url: str):
    """Gestión y consulta de facturas"""
    
    st.subheader("📋 Gestión de Facturas")
    
    # Filtros de búsqueda
    col1, col2, col3 = st.columns(3)
    
    with col1:
        estado_filtro = st.selectbox(
            "Estado:",
            ["Todas", "Pendiente", "Pagada", "Vencida", "Anulada"]
        )
    
    with col2:
        fecha_desde = st.date_input("Desde:", value=None)
    
    with col3:
        fecha_hasta = st.date_input("Hasta:", value=None)
    
    # Filtros adicionales
    col1, col2 = st.columns(2)
    
    with col1:
        try:
            response_clientes = requests.get(f"{backend_url}/api/facturacion/clientes")
            clientes = response_clientes.json() if response_clientes.status_code == 200 else []
            
            opciones_clientes = ["Todos los clientes"] + [
                f"{c['codigo_cliente']} - {c['nombre']}"
                for c in clientes
            ]
            cliente_filtro = st.selectbox("Cliente:", opciones_clientes)
        except:
            cliente_filtro = "Todos los clientes"
    
    with col2:
        numero_factura = st.text_input("Número de factura:", help="Buscar por número específico")
    
    if st.button("🔍 Buscar Facturas", use_container_width=True):
        # Guardar búsqueda en session_state
        buscar_facturas(
            backend_url,
            estado_filtro,
            fecha_desde,
            fecha_hasta,
            cliente_filtro,
            numero_factura
        )
    
    # Mostrar facturas si existen en session_state
    if 'facturas_encontradas' in st.session_state and st.session_state.facturas_encontradas:
        mostrar_facturas(st.session_state.facturas_encontradas, backend_url)

def buscar_facturas(
    backend_url: str,
    estado_filtro: str,
    fecha_desde: date,
    fecha_hasta: date,
    cliente_filtro: str,
    numero_factura: str
):
    """Buscar y mostrar facturas"""
    
    try:
        params = {}
        
        if estado_filtro != "Todas":
            params["estado"] = estado_filtro
        
        if fecha_desde:
            params["fecha_desde"] = fecha_desde.isoformat()
        
        if fecha_hasta:
            params["fecha_hasta"] = fecha_hasta.isoformat()
        
        if cliente_filtro != "Todos los clientes":
            codigo_cliente = cliente_filtro.split(" - ")[0]
            params["codigo_cliente"] = codigo_cliente
        
        if numero_factura:
            params["numero_factura"] = numero_factura
        
        with st.spinner("Buscando facturas..."):
            response = requests.get(f"{backend_url}/api/facturacion/facturas", params=params)
        
        if response.status_code == 200:
            facturas = response.json()
            
            if facturas:
                # Guardar en session_state
                st.session_state.facturas_encontradas = facturas
                st.rerun()
            else:
                st.info("📭 No se encontraron facturas con los criterios especificados")
                st.session_state.facturas_encontradas = None
        else:
            st.error(f"Error al buscar facturas: {response.status_code}")
            st.session_state.facturas_encontradas = None
            
    except Exception as e:
        st.error(f"Error al buscar facturas: {e}")
        st.session_state.facturas_encontradas = None

def mostrar_facturas(facturas: List[Dict], backend_url: str):
    """Mostrar lista de facturas"""
    
    st.markdown(f"### 📋 Facturas Encontradas ({len(facturas)})")
    
    # Métricas resumen
    total_facturas = len(facturas)
    total_valor = sum(float(f.get('total', 0)) for f in facturas)
    facturas_pendientes = len([f for f in facturas if f.get('estado_factura') == 'EMITIDA'])
    facturas_pagadas = len([f for f in facturas if f.get('estado_factura') == 'PAGADA'])
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Facturas", total_facturas)
    
    with col2:
        st.metric("Valor Total", f"${total_valor:,.2f}")
    
    with col3:
        st.metric("Pendientes", facturas_pendientes)
    
    with col4:
        st.metric("Pagadas", facturas_pagadas)
    
    st.markdown("---")
    
    # Tabla de facturas
    df_facturas = pd.DataFrame(facturas)
    
    # Formatear para visualización
    df_display = df_facturas.copy()
    
    # Formatear fechas y valores
    if 'fecha_emision' in df_display.columns:
        df_display['fecha_factura'] = pd.to_datetime(df_display['fecha_emision']).dt.strftime('%d/%m/%Y')
    if 'fecha_vencimiento' in df_display.columns:
        df_display['fecha_vencimiento'] = pd.to_datetime(df_display['fecha_vencimiento']).dt.strftime('%d/%m/%Y')
    
    for col in ['subtotal', 'impuesto_iva', 'total']:
        if col in df_display.columns:
            df_display[f'{col}_fmt'] = df_display[col].apply(lambda x: f"${float(x):,.2f}")
    
    # Seleccionar columnas para mostrar
    columnas_mostrar = ['numero_factura', 'fecha_factura', 'estado_factura', 'total_fmt']
    nombres_columnas = ['Número', 'Fecha', 'Estado', 'Total']
    
    # Verificar que las columnas existan
    columnas_disponibles = [col for col in columnas_mostrar if col in df_display.columns]
    
    if columnas_disponibles:
        df_final = df_display[columnas_disponibles].copy()
        df_final.columns = nombres_columnas[:len(columnas_disponibles)]
        
        # Mostrar tabla
        st.dataframe(
            df_final,
            use_container_width=True,
            hide_index=True
        )
        
        st.markdown("---")
        st.markdown("### 🎯 Seleccionar Factura para Acciones")
        
        # Selector de factura
        opciones_facturas = [
            f"{f.get('numero_factura', 'N/A')} - {f.get('estado_factura', 'N/A')} - ${float(f.get('total', 0)):,.2f}"
            for f in facturas
        ]
        
        factura_seleccionada_idx = st.selectbox(
            "Selecciona una factura:",
            range(len(opciones_facturas)),
            format_func=lambda x: opciones_facturas[x],
            key="selector_factura_gestion"
        )
        
        if factura_seleccionada_idx is not None:
            factura_seleccionada = facturas[factura_seleccionada_idx]
            
            st.markdown("---")
            
            # Mostrar información de la factura seleccionada
            col1, col2, col3 = st.columns(3)
            with col1:
                st.info(f"**Factura:** {factura_seleccionada.get('numero_factura', 'N/A')}")
            with col2:
                st.info(f"**Estado:** {factura_seleccionada.get('estado_factura', 'N/A')}")
            with col3:
                st.info(f"**Total:** ${float(factura_seleccionada.get('total', 0)):,.2f}")
            
            st.markdown("### 🔧 Acciones Disponibles")
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                if st.button("👁️ Ver Detalles", use_container_width=True):
                    st.session_state.mostrar_detalle = True
            
            with col2:
                if st.button("💰 Marcar como Pagada", use_container_width=True):
                    marcar_como_pagada(backend_url, factura_seleccionada['id_factura'])
            
            with col3:
                if st.button("📥 Descargar", use_container_width=True, type="primary"):
                    st.session_state.mostrar_descarga = True
            
            with col4:
                if st.button("❌ Anular", use_container_width=True):
                    st.session_state.mostrar_anular = True
            
            # Mostrar secciones según botones presionados
            if st.session_state.get('mostrar_detalle', False):
                mostrar_detalle_factura(factura_seleccionada, backend_url)
                if st.button("🔙 Cerrar Detalles"):
                    st.session_state.mostrar_detalle = False
                    st.rerun()
            
            if st.session_state.get('mostrar_descarga', False):
                mostrar_opciones_descarga(factura_seleccionada, backend_url)
                if st.button("🔙 Cerrar Descarga"):
                    st.session_state.mostrar_descarga = False
                    st.rerun()
            
            if st.session_state.get('mostrar_anular', False):
                st.warning("⚠️ ¿Estás seguro de que deseas anular esta factura?")
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("✅ Sí, Anular", type="primary", use_container_width=True):
                        anular_factura(backend_url, factura_seleccionada['id_factura'])
                        st.session_state.mostrar_anular = False
                with col2:
                    if st.button("❌ Cancelar", use_container_width=True):
                        st.session_state.mostrar_anular = False
                        st.rerun()

def mostrar_detalle_factura(factura: Dict, backend_url: str):
    """Mostrar detalle completo de una factura"""
    
    # Obtener detalle completo de la factura desde el backend
    try:
        response = requests.get(f"{backend_url}/api/facturacion/facturas/{factura['id_factura']}")
        if response.status_code == 200:
            factura_completa = response.json()
        else:
            factura_completa = factura
    except:
        factura_completa = factura
    
    with st.expander(f"📄 Detalle Factura {factura_completa.get('numero_factura', 'N/A')}", expanded=True):
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**📋 Información General:**")
            st.text(f"Número: {factura_completa.get('numero_factura', 'N/A')}")
            st.text(f"Fecha: {factura_completa.get('fecha_emision', 'N/A')[:10]}")
            st.text(f"Vencimiento: {factura_completa.get('fecha_vencimiento', 'N/A')[:10]}")
            st.text(f"Estado: {factura_completa.get('estado_factura', 'N/A')}")
            st.text(f"Tipo: {factura_completa.get('tipo_factura', 'N/A')}")
        
        with col2:
            st.markdown("**💰 Totales:**")
            st.text(f"Subtotal: ${float(factura_completa.get('subtotal', 0)):,.2f}")
            st.text(f"IVA: ${float(factura_completa.get('impuesto_iva', 0)):,.2f}")
            st.text(f"Total: ${float(factura_completa.get('total', 0)):,.2f}")
        
        # Cliente
        st.markdown("**👤 Cliente:**")
        cliente_info = factura_completa.get('cliente', {})
        if isinstance(cliente_info, dict):
            st.text(f"Nombre: {cliente_info.get('nombre', 'N/A')}")
            st.text(f"NIT/CC: {cliente_info.get('nit', 'N/A')}")
            st.text(f"Dirección: {cliente_info.get('direccion', 'N/A')}")
        
        # Items de la factura
        if 'detalles' in factura_completa and factura_completa['detalles']:
            st.markdown("**📦 Items:**")
            
            df_items = pd.DataFrame(factura_completa['detalles'])
            df_items_display = df_items.copy()
            
            # Formatear columnas monetarias
            for col in ['precio_unitario', 'subtotal']:
                if col in df_items_display.columns:
                    df_items_display[col] = df_items_display[col].apply(lambda x: f"${float(x):,.2f}")
            
            # Seleccionar columnas relevantes
            cols_to_show = []
            col_names = []
            
            if 'nombre_producto' in df_items_display.columns:
                cols_to_show.append('nombre_producto')
                col_names.append('Producto')
            if 'cantidad' in df_items_display.columns:
                cols_to_show.append('cantidad')
                col_names.append('Cantidad')
            if 'precio_unitario' in df_items_display.columns:
                cols_to_show.append('precio_unitario')
                col_names.append('Precio Unit.')
            if 'subtotal' in df_items_display.columns:
                cols_to_show.append('subtotal')
                col_names.append('Subtotal')
            
            if cols_to_show:
                df_final = df_items_display[cols_to_show].copy()
                df_final.columns = col_names
                st.dataframe(df_final, use_container_width=True, hide_index=True)


def mostrar_opciones_descarga(factura: Dict, backend_url: str):
    """Mostrar opciones de descarga de factura"""
    
    st.markdown("---")
    st.markdown("### 📥 Descargar Factura")
    
    # Obtener detalle completo de la factura con datos del cliente y productos
    try:
        response = requests.get(f"{backend_url}/api/facturacion/facturas/{factura['id_factura']}/completa")
        if response.status_code == 200:
            factura_completa = response.json()
        else:
            factura_completa = factura
    except Exception as e:
        st.warning(f"⚠️ No se pudieron cargar los datos completos: {e}")
        factura_completa = factura
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("📄 Descargar PDF", use_container_width=True, type="primary"):
            generar_pdf_factura(factura_completa, backend_url)
    
    with col2:
        if st.button("📊 Descargar Excel", use_container_width=True):
            generar_excel_factura(factura_completa, backend_url)
    
    with col3:
        if st.button("📋 Descargar JSON", use_container_width=True):
            generar_json_factura(factura_completa, backend_url)


def generar_pdf_factura(factura: Dict, backend_url: str):
    """Generar factura en formato PDF profesional tipo DIAN"""
    
    try:
        # Obtener configuración de empresa e IVA
        iva_porcentaje = 19.0  # Valor por defecto
        empresa_nombre = "EMPRESA"
        empresa_nit = "N/A"
        empresa_direccion = "N/A"
        empresa_telefono = "N/A"
        empresa_email = "N/A"
        try:
            response_config = requests.get(f"{backend_url}/api/facturacion/configuracion")
            if response_config.status_code == 200:
                config = response_config.json()
                iva_porcentaje = float(config.get('iva_porcentaje', 19.0))
                empresa_nombre = config.get('empresa_nombre', 'EMPRESA')
                empresa_nit = config.get('empresa_nit', 'N/A')
                empresa_direccion = config.get('empresa_direccion', 'N/A')
                empresa_telefono = config.get('empresa_telefono', 'N/A')
                empresa_email = config.get('empresa_email', 'N/A')
        except:
            pass
        
        with st.spinner("📄 Generando PDF..."):
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, 
                                   rightMargin=0.6*inch, leftMargin=0.6*inch,
                                   topMargin=0.4*inch, bottomMargin=0.4*inch)
            
            # Contenedor de elementos
            elements = []
            styles = getSampleStyleSheet()
            
            # Estilos personalizados
            company_style = ParagraphStyle(
                'CompanyName',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=4,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            title_style = ParagraphStyle(
                'InvoiceTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.white,
                spaceAfter=0,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            header_style = ParagraphStyle(
                'SectionHeader',
                parent=styles['Heading2'],
                fontSize=11,
                textColor=colors.HexColor('#2c3e50'),
                spaceAfter=8,
                fontName='Helvetica-Bold'
            )
            
            normal_style = ParagraphStyle(
                'NormalText',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#333333'),
                leading=12
            )
            
            small_style = ParagraphStyle(
                'SmallText',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#666666'),
                leading=10
            )
            
            # Encabezado con nombre empresa
            empresa_nombre_p = Paragraph(f"<b>{empresa_nombre.upper()}</b>", company_style)
            elements.append(empresa_nombre_p)
            
            empresa_info = Paragraph(f"NIT: {empresa_nit}", small_style)
            elements.append(empresa_info)
            
            empresa_contacto = Paragraph(f"{empresa_direccion} | Tel: {empresa_telefono} | {empresa_email}", small_style)
            elements.append(empresa_contacto)
            
            elements.append(Spacer(1, 0.15*inch))
            
            # Rectángulo con título FACTURA
            numero_factura = factura.get('numero_factura', 'N/A')
            estado = factura.get('estado', 'N/A')
            
            # Tabla de título con color
            titulo_data = [[Paragraph("FACTURA ELECTRÓNICA DE VENTA", title_style)]]
            tabla_titulo = Table(titulo_data, colWidths=[7*inch])
            tabla_titulo.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
                ('TOPPADDING', (0, 0), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))
            elements.append(tabla_titulo)
            
            elements.append(Spacer(1, 0.1*inch))
            
            # Información de la factura y cliente en dos columnas
            fecha_emision = factura.get('fecha_emision', 'N/A')[:10] if factura.get('fecha_emision') else 'N/A'
            fecha_venc = factura.get('fecha_vencimiento', 'N/A')[:10] if factura.get('fecha_vencimiento') else 'N/A'
            
            cliente_info = factura.get('cliente', {})
            cliente_nombre = cliente_info.get('nombre', 'N/A') if isinstance(cliente_info, dict) else 'N/A'
            cliente_nit = cliente_info.get('nit', 'N/A') if isinstance(cliente_info, dict) else 'N/A'
            cliente_dir = cliente_info.get('direccion', 'N/A') if isinstance(cliente_info, dict) else 'N/A'
            cliente_tel = cliente_info.get('telefono_principal', 'N/A') if isinstance(cliente_info, dict) else 'N/A'
            cliente_email = cliente_info.get('email', 'N/A') if isinstance(cliente_info, dict) else 'N/A'
            
            info_factura_cliente = [
                [
                    Paragraph("<b>DATOS DE LA FACTURA</b>", header_style),
                    Paragraph("<b>DATOS DEL CLIENTE</b>", header_style)
                ],
                [
                    Paragraph(f"<b>No. Factura:</b> {numero_factura}", normal_style),
                    Paragraph(f"<b>Razón Social:</b> {cliente_nombre}", normal_style)
                ],
                [
                    Paragraph(f"<b>Fecha Emisión:</b> {fecha_emision}", normal_style),
                    Paragraph(f"<b>NIT/CC:</b> {cliente_nit}", normal_style)
                ],
                [
                    Paragraph(f"<b>Fecha Vencimiento:</b> {fecha_venc}", normal_style),
                    Paragraph(f"<b>Dirección:</b> {cliente_dir}", normal_style)
                ],
                [
                    Paragraph(f"<b>Estado:</b> {estado.upper()}", normal_style),
                    Paragraph(f"<b>Teléfono:</b> {cliente_tel}", normal_style)
                ],
                [
                    "",
                    Paragraph(f"<b>Email:</b> {cliente_email}", normal_style)
                ]
            ]
            
            tabla_info = Table(info_factura_cliente, colWidths=[3.5*inch, 3.5*inch])
            tabla_info.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#e8f4f8')),
                ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#e8f4f8')),
                ('BOX', (0, 0), (0, -1), 1, colors.HexColor('#3498db')),
                ('BOX', (1, 0), (1, -1), 1, colors.HexColor('#3498db')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            elements.append(tabla_info)
            elements.append(Spacer(1, 0.2*inch))
            
            # Detalle de productos con numeración
            detalle_header = Paragraph("DETALLE DE LA OPERACIÓN", header_style)
            elements.append(detalle_header)
            
            detalles = factura.get('detalles', [])
            if detalles:
                data_items = [[
                    Paragraph('<b>Item</b>', normal_style),
                    Paragraph('<b>Código</b>', normal_style),
                    Paragraph('<b>Descripción</b>', normal_style),
                    Paragraph('<b>Cant.</b>', normal_style),
                    Paragraph('<b>Vlr. Unitario</b>', normal_style),
                    Paragraph('<b>Vlr. Total</b>', normal_style)
                ]]
                
                for idx, item in enumerate(detalles, 1):
                    codigo_producto = item.get('codigo_producto', 'N/A')
                    nombre = item.get('nombre_producto', 'Producto')
                    cantidad = item.get('cantidad', 0)
                    precio_unit = float(item.get('precio_unitario', 0))
                    subtotal = float(item.get('subtotal_linea', 0))
                    
                    data_items.append([
                        Paragraph(f"{idx}", normal_style),
                        Paragraph(codigo_producto, normal_style),
                        Paragraph(nombre, normal_style),
                        Paragraph(f"{cantidad}", normal_style),
                        Paragraph(f"${precio_unit:,.2f}", normal_style),
                        Paragraph(f"${subtotal:,.2f}", normal_style)
                    ])
                
                tabla_items = Table(data_items, colWidths=[0.4*inch, 0.8*inch, 2.8*inch, 0.6*inch, 1.2*inch, 1.2*inch])
                tabla_items.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('TOPPADDING', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                    ('ALIGN', (0, 1), (1, -1), 'CENTER'),
                    ('ALIGN', (2, 1), (2, -1), 'LEFT'),
                    ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
                    ('TOPPADDING', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ]))
                
                elements.append(tabla_items)
                elements.append(Spacer(1, 0.15*inch))
            
            # Observaciones y Totales lado a lado
            observaciones = factura.get('observaciones') or 'Sin observaciones'
            
            obs_totales_data = [
                [
                    Paragraph("<b>OBSERVACIONES:</b>", normal_style),
                    ""
                ],
                [
                    Paragraph(str(observaciones), small_style),
                    ""
                ]
            ]
            
            # Calcular totales
            subtotal = float(factura.get('subtotal', 0))
            iva = float(factura.get('impuesto_iva', 0))
            total = float(factura.get('total', 0))
            
            # Tabla de totales separada
            totales_data = [
                [Paragraph('<b>Subtotal:</b>', normal_style), Paragraph(f"${subtotal:,.2f}", normal_style)],
                [Paragraph(f'<b>IVA ({iva_porcentaje:.0f}%):</b>', normal_style), Paragraph(f"${iva:,.2f}", normal_style)],
                [Paragraph('<b>TOTAL A PAGAR:</b>', header_style), Paragraph(f"<b>${total:,.2f}</b>", header_style)]
            ]
            
            tabla_totales = Table(totales_data, colWidths=[2*inch, 1.5*inch])
            tabla_totales.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('LINEABOVE', (0, 2), (-1, 2), 2, colors.HexColor('#2c3e50')),
                ('LINEBELOW', (0, 2), (-1, 2), 2, colors.HexColor('#2c3e50')),
                ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#e8f4f8')),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ]))
            
            # Tabla que combina observaciones y totales
            contenedor = [[
                Paragraph(f"<b>OBSERVACIONES:</b><br/>{str(observaciones)}", normal_style),
                tabla_totales
            ]]
            
            tabla_contenedor = Table(contenedor, colWidths=[3.5*inch, 3.5*inch])
            tabla_contenedor.setStyle(TableStyle([
                ('VALIGN', (0, 0), (0, 0), 'TOP'),
                ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (0, 0), 0),
                ('RIGHTPADDING', (1, 0), (1, 0), 0),
            ]))
            
            elements.append(tabla_contenedor)
            elements.append(Spacer(1, 0.3*inch))
            
            # Información legal y pie de página
            legal_style = ParagraphStyle(
                'Legal',
                parent=styles['Normal'],
                fontSize=7,
                textColor=colors.HexColor('#7f8c8d'),
                alignment=TA_CENTER,
                leading=9
            )
            
            cufe = "CUFE: " + "A1B2C3D4E5F6G7H8I9J0K1L2M3N4O5P6Q7R8S9T0U1V2W3X4Y5Z6"
            
            legal_info = f"""
            <b>INFORMACIÓN TRIBUTARIA</b><br/>
            {cufe}<br/>
            Factura electrónica generada de acuerdo a la Resolución DIAN 000042 del 05 de mayo de 2020<br/>
            Esta factura es un título valor según Art. 772 del Código de Comercio<br/>
            <b>Gracias por su compra - Conserve este documento para efectos tributarios</b>
            """
            
            pie = Paragraph(legal_info, legal_style)
            elements.append(pie)
            
            # Construir PDF
            doc.build(elements)
            
            # Preparar descarga
            buffer.seek(0)
            nombre_archivo = f"FV_{numero_factura.replace('/', '-')}_{fecha_emision.replace('-', '')}.pdf"
            
            st.download_button(
                label="📥 Descargar PDF",
                data=buffer,
                file_name=nombre_archivo,
                mime="application/pdf",
                type="primary",
                use_container_width=True
            )
            
            st.success("✅ Factura PDF generada exitosamente")
            
    except Exception as e:
        st.error(f"❌ Error al generar PDF: {str(e)}")
        st.exception(e)


def generar_excel_factura(factura: Dict, backend_url: str):
    """Generar factura en formato Excel profesional tipo DIAN"""
    
    try:
        # Obtener configuración de empresa e IVA
        iva_porcentaje = 19.0  # Valor por defecto
        empresa_nombre = "EMPRESA"
        empresa_nit = "N/A"
        empresa_direccion = "N/A"
        empresa_telefono = "N/A"
        empresa_email = "N/A"
        try:
            response_config = requests.get(f"{backend_url}/api/facturacion/configuracion")
            if response_config.status_code == 200:
                config = response_config.json()
                iva_porcentaje = float(config.get('iva_porcentaje', 19.0))
                empresa_nombre = config.get('empresa_nombre', 'EMPRESA')
                empresa_nit = config.get('empresa_nit', 'N/A')
                empresa_direccion = config.get('empresa_direccion', 'N/A')
                empresa_telefono = config.get('empresa_telefono', 'N/A')
                empresa_email = config.get('empresa_email', 'N/A')
        except:
            pass
        
        with st.spinner("📊 Generando Excel..."):
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                workbook = writer.book
                
                # === HOJA 1: FACTURA COMPLETA ===
                ws_factura = workbook.create_sheet('FACTURA', 0)
                
                # Encabezado empresa
                ws_factura['A1'] = empresa_nombre.upper()
                ws_factura['A1'].font = Font(name='Arial', size=16, bold=True)
                
                ws_factura['A2'] = f'NIT: {empresa_nit}'
                ws_factura['A2'].font = Font(name='Arial', size=9)
                
                ws_factura['A3'] = f'{empresa_direccion} | Tel: {empresa_telefono} | {empresa_email}'
                ws_factura['A3'].font = Font(name='Arial', size=9, color='666666')
                
                # Título factura
                ws_factura.merge_cells('A5:F5')
                ws_factura['A5'] = 'FACTURA ELECTRÓNICA DE VENTA'
                ws_factura['A5'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
                ws_factura['A5'].fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
                ws_factura['A5'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Datos de la factura
                row = 7
                ws_factura[f'A{row}'] = 'DATOS DE LA FACTURA'
                ws_factura[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
                ws_factura[f'A{row}'].fill = PatternFill(start_color='e8f4f8', end_color='e8f4f8', fill_type='solid')
                ws_factura.merge_cells(f'A{row}:C{row}')
                
                ws_factura[f'D{row}'] = 'DATOS DEL CLIENTE'
                ws_factura[f'D{row}'].font = Font(name='Arial', size=11, bold=True)
                ws_factura[f'D{row}'].fill = PatternFill(start_color='e8f4f8', end_color='e8f4f8', fill_type='solid')
                ws_factura.merge_cells(f'D{row}:F{row}')
                
                cliente_info = factura.get('cliente', {})
                
                datos_factura_cliente = [
                    ('No. Factura:', factura.get('numero_factura', 'N/A'), 'Razón Social:', cliente_info.get('nombre', 'N/A') if isinstance(cliente_info, dict) else 'N/A'),
                    ('Fecha Emisión:', factura.get('fecha_emision', 'N/A')[:10], 'NIT/CC:', cliente_info.get('nit', 'N/A') if isinstance(cliente_info, dict) else 'N/A'),
                    ('Fecha Vencimiento:', factura.get('fecha_vencimiento', 'N/A')[:10], 'Dirección:', cliente_info.get('direccion', 'N/A') if isinstance(cliente_info, dict) else 'N/A'),
                    ('Estado:', factura.get('estado', 'N/A').upper(), 'Teléfono:', cliente_info.get('telefono_principal', 'N/A') if isinstance(cliente_info, dict) else 'N/A'),
                    ('', '', 'Email:', cliente_info.get('email', 'N/A') if isinstance(cliente_info, dict) else 'N/A'),
                ]
                
                row += 1
                for campo_fac, valor_fac, campo_cli, valor_cli in datos_factura_cliente:
                    ws_factura[f'A{row}'] = campo_fac
                    ws_factura[f'A{row}'].font = Font(name='Arial', size=9, bold=True)
                    ws_factura[f'B{row}'] = valor_fac
                    ws_factura[f'B{row}'].font = Font(name='Arial', size=9)
                    
                    ws_factura[f'D{row}'] = campo_cli
                    ws_factura[f'D{row}'].font = Font(name='Arial', size=9, bold=True)
                    ws_factura[f'E{row}'] = valor_cli
                    ws_factura[f'E{row}'].font = Font(name='Arial', size=9)
                    ws_factura.merge_cells(f'E{row}:F{row}')
                    row += 1
                
                # Detalle de productos
                row += 2
                ws_factura[f'A{row}'] = 'DETALLE DE LA OPERACIÓN'
                ws_factura[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
                ws_factura.merge_cells(f'A{row}:F{row}')
                
                row += 1
                headers = ['Item', 'Código', 'Descripción', 'Cant.', 'Vlr. Unitario', 'Vlr. Total']
                for col, header in enumerate(headers, 1):
                    cell = ws_factura.cell(row=row, column=col, value=header)
                    cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='34495e', end_color='34495e', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                detalles = factura.get('detalles', [])
                for idx, item in enumerate(detalles, 1):
                    row += 1
                    ws_factura[f'A{row}'] = idx
                    ws_factura[f'B{row}'] = item.get('codigo_producto', 'N/A')
                    ws_factura[f'C{row}'] = item.get('nombre_producto', 'N/A')
                    ws_factura[f'D{row}'] = item.get('cantidad', 0)
                    ws_factura[f'E{row}'] = float(item.get('precio_unitario', 0))
                    ws_factura[f'F{row}'] = float(item.get('subtotal_linea', 0))
                    
                    # Formato moneda
                    ws_factura[f'E{row}'].number_format = '$#,##0.00'
                    ws_factura[f'F{row}'].number_format = '$#,##0.00'
                    ws_factura[f'D{row}'].alignment = Alignment(horizontal='center')
                    
                    # Alternar colores
                    if idx % 2 == 0:
                        for col in range(1, 7):
                            ws_factura.cell(row=row, column=col).fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
                
                # Observaciones y totales
                row += 2
                ws_factura[f'A{row}'] = 'OBSERVACIONES:'
                ws_factura[f'A{row}'].font = Font(name='Arial', size=9, bold=True)
                ws_factura.merge_cells(f'A{row}:C{row}')
                
                row += 1
                ws_factura[f'A{row}'] = factura.get('observaciones') or 'Sin observaciones'
                ws_factura[f'A{row}'].font = Font(name='Arial', size=9)
                ws_factura.merge_cells(f'A{row}:C{row+2}')
                ws_factura[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                
                # Totales
                row_totales = row
                ws_factura[f'E{row_totales}'] = 'Subtotal:'
                ws_factura[f'E{row_totales}'].font = Font(name='Arial', size=9, bold=True)
                ws_factura[f'E{row_totales}'].alignment = Alignment(horizontal='right')
                ws_factura[f'F{row_totales}'] = float(factura.get('subtotal', 0))
                ws_factura[f'F{row_totales}'].number_format = '$#,##0.00'
                
                row_totales += 1
                ws_factura[f'E{row_totales}'] = f'IVA ({iva_porcentaje:.0f}%):'
                ws_factura[f'E{row_totales}'].font = Font(name='Arial', size=9, bold=True)
                ws_factura[f'E{row_totales}'].alignment = Alignment(horizontal='right')
                ws_factura[f'F{row_totales}'] = float(factura.get('impuesto_iva', 0))
                ws_factura[f'F{row_totales}'].number_format = '$#,##0.00'
                
                row_totales += 1
                ws_factura[f'E{row_totales}'] = 'TOTAL A PAGAR:'
                ws_factura[f'E{row_totales}'].font = Font(name='Arial', size=11, bold=True)
                ws_factura[f'E{row_totales}'].alignment = Alignment(horizontal='right')
                ws_factura[f'E{row_totales}'].fill = PatternFill(start_color='e8f4f8', end_color='e8f4f8', fill_type='solid')
                ws_factura[f'F{row_totales}'] = float(factura.get('total', 0))
                ws_factura[f'F{row_totales}'].number_format = '$#,##0.00'
                ws_factura[f'F{row_totales}'].font = Font(name='Arial', size=11, bold=True)
                ws_factura[f'F{row_totales}'].fill = PatternFill(start_color='e8f4f8', end_color='e8f4f8', fill_type='solid')
                
                # Pie de página legal
                row += 5
                ws_factura.merge_cells(f'A{row}:F{row+2}')
                legal_text = (
                    'INFORMACIÓN TRIBUTARIA\n'
                    'CUFE: A1B2C3D4E5F6G7H8I9J0K1L2M3N4O5P6Q7R8S9T0U1V2W3X4Y5Z6\n'
                    'Factura electrónica generada de acuerdo a la Resolución DIAN 000042 del 05 de mayo de 2020\n'
                    'Esta factura es un título valor según Art. 772 del Código de Comercio'
                )
                ws_factura[f'A{row}'] = legal_text
                ws_factura[f'A{row}'].font = Font(name='Arial', size=7, color='7f8c8d')
                ws_factura[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Ajustar anchos de columna
                ws_factura.column_dimensions['A'].width = 8
                ws_factura.column_dimensions['B'].width = 15
                ws_factura.column_dimensions['C'].width = 30
                ws_factura.column_dimensions['D'].width = 10
                ws_factura.column_dimensions['E'].width = 18
                ws_factura.column_dimensions['F'].width = 18
                
                # === HOJA 2: RESUMEN EJECUTIVO ===
                resumen_data = {
                    '': [
                        'NÚMERO DE FACTURA',
                        'FECHA DE EMISIÓN',
                        'CLIENTE',
                        'NIT/CC',
                        '',
                        'SUBTOTAL',
                        f'IVA ({iva_porcentaje:.0f}%)',
                        'TOTAL A PAGAR',
                        '',
                        'ESTADO',
                        'ITEMS',
                        'FECHA VENCIMIENTO'
                    ],
                    'INFORMACIÓN': [
                        factura.get('numero_factura', 'N/A'),
                        factura.get('fecha_emision', 'N/A')[:10],
                        cliente_info.get('nombre', 'N/A') if isinstance(cliente_info, dict) else 'N/A',
                        cliente_info.get('nit', 'N/A') if isinstance(cliente_info, dict) else 'N/A',
                        '',
                        float(factura.get('subtotal', 0)),
                        float(factura.get('impuesto_iva', 0)),
                        float(factura.get('total', 0)),
                        '',
                        factura.get('estado', 'N/A').upper(),
                        len(detalles),
                        factura.get('fecha_vencimiento', 'N/A')[:10]
                    ]
                }
                
                df_resumen = pd.DataFrame(resumen_data)
                df_resumen.to_excel(writer, sheet_name='RESUMEN', index=False)
                
                ws_resumen = writer.sheets['RESUMEN']
                for row in ws_resumen.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = Font(bold=True, size=12, color='FFFFFF')
                        cell.fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
                
                # Formatear montos en resumen
                ws_resumen['B7'].number_format = '$#,##0.00'
                ws_resumen['B8'].number_format = '$#,##0.00'
                ws_resumen['B9'].number_format = '$#,##0.00'
                ws_resumen['B9'].font = Font(bold=True, size=12)
                
                ws_resumen.column_dimensions['A'].width = 25
                ws_resumen.column_dimensions['B'].width = 30
                
            # Eliminar hoja por defecto
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            output.seek(0)
            numero_factura = factura.get('numero_factura', 'N/A').replace('/', '-')
            fecha = factura.get('fecha_emision', 'N/A')[:10].replace('-', '')
            nombre_archivo = f"FV_{numero_factura}_{fecha}.xlsx"
            
            st.download_button(
                label="📥 Descargar Excel",
                data=output,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
            
            st.success("✅ Factura Excel generada exitosamente")
            
    except Exception as e:
        st.error(f"❌ Error al generar Excel: {str(e)}")
        st.exception(e)


def generar_json_factura(factura: Dict, backend_url: str):
    """Generar factura en formato JSON estructurado compatible DIAN"""
    
    try:
        # Obtener configuración de empresa e IVA
        iva_porcentaje = 19.0  # Valor por defecto
        empresa_nombre = "EMPRESA"
        empresa_nit = "N/A"
        empresa_direccion = "N/A"
        empresa_telefono = "N/A"
        empresa_email = "N/A"
        empresa_web = ""
        try:
            response_config = requests.get(f"{backend_url}/api/facturacion/configuracion")
            if response_config.status_code == 200:
                config = response_config.json()
                iva_porcentaje = float(config.get('iva_porcentaje', 19.0))
                empresa_nombre = config.get('empresa_nombre', 'EMPRESA')
                empresa_nit = config.get('empresa_nit', 'N/A')
                empresa_direccion = config.get('empresa_direccion', 'N/A')
                empresa_telefono = config.get('empresa_telefono', 'N/A')
                empresa_email = config.get('empresa_email', 'N/A')
                empresa_web = config.get('empresa_web', '')
        except:
            pass
        
        with st.spinner("📋 Generando JSON..."):
            
            # Obtener información del cliente
            cliente_info = factura.get('cliente', {})
            if not isinstance(cliente_info, dict):
                cliente_info = {}
            
            # Calcular totales
            detalles = factura.get('detalles', [])
            subtotal = float(factura.get('subtotal', 0))
            iva = float(factura.get('impuesto_iva', 0))
            total = float(factura.get('total', 0))
            
            # Estructurar según formato DIAN
            factura_json = {
                "documento": {
                    "tipo": "FACTURA_ELECTRONICA",
                    "version": "1.0",
                    "fecha_generacion": datetime.now().isoformat(),
                    "ambiente": "PRODUCCION"
                },
                "emisor": {
                    "razon_social": empresa_nombre,
                    "nit": empresa_nit,
                    "regimen": "COMUN",
                    "responsabilidades_fiscales": ["O-13", "R-99-PN"],
                    "direccion": {
                        "calle": empresa_direccion,
                        "ciudad": "N/A",
                        "departamento": "N/A",
                        "codigo_postal": "N/A",
                        "pais": "SV"
                    },
                    "contacto": {
                        "telefono": empresa_telefono,
                        "email": empresa_email,
                        "sitio_web": empresa_web
                    }
                },
                "factura": {
                    "numero": factura.get('numero_factura', 'N/A'),
                    "prefijo": "FV",
                    "consecutivo": factura.get('numero_factura', 'N/A').replace('FV-', ''),
                    "fecha_emision": factura.get('fecha_emision', 'N/A'),
                    "hora_emision": datetime.now().strftime("%H:%M:%S"),
                    "fecha_vencimiento": factura.get('fecha_vencimiento', 'N/A'),
                    "estado": factura.get('estado', 'N/A').upper(),
                    "tipo_operacion": "VENTA_NACIONAL",
                    "forma_pago": "CONTADO",
                    "medio_pago": "EFECTIVO",
                    "moneda": "COP"
                },
                "adquiriente": {
                    "tipo_persona": "JURIDICA",
                    "tipo_documento": "NIT",
                    "numero_documento": cliente_info.get('nit', 'N/A'),
                    "razon_social": cliente_info.get('nombre', 'N/A'),
                    "nombre_comercial": cliente_info.get('nombre', 'N/A'),
                    "regimen": "SIMPLIFICADO",
                    "responsabilidades_fiscales": ["R-99-PN"],
                    "direccion": {
                        "calle": cliente_info.get('direccion', 'N/A'),
                        "ciudad": "N/A",
                        "departamento": "N/A",
                        "pais": "CO"
                    },
                    "contacto": {
                        "telefono": cliente_info.get('telefono_principal', 'N/A'),
                        "email": cliente_info.get('email', 'N/A')
                    }
                },
                "items": [
                    {
                        "numero_linea": idx,
                        "codigo_producto": item.get('codigo_producto', 'N/A'),
                        "codigo_estandar": f"999{item.get('producto_id', '000')}",
                        "descripcion": item.get('nombre_producto', 'N/A'),
                        "cantidad": float(item.get('cantidad', 0)),
                        "unidad_medida": "UND",
                        "precio_unitario": float(item.get('precio_unitario', 0)),
                        "valor_bruto": float(item.get('subtotal_linea', 0)),
                        "descuento": 0.0,
                        "cargo": 0.0,
                        "valor_neto": float(item.get('subtotal_linea', 0)),
                        "impuestos": [
                            {
                                "tipo": "IVA",
                                "base": float(item.get('subtotal_linea', 0)),
                                "porcentaje": iva_porcentaje,
                                "valor": float(item.get('subtotal_linea', 0)) * (iva_porcentaje / 100)
                            }
                        ]
                    }
                    for idx, item in enumerate(detalles, 1)
                ],
                "totales": {
                    "cantidad_lineas": len(detalles),
                    "subtotal": subtotal,
                    "descuentos": 0.0,
                    "cargos": 0.0,
                    "base_imponible": subtotal,
                    "impuestos": [
                        {
                            "tipo": "IVA",
                            "base": subtotal,
                            "porcentaje": iva_porcentaje,
                            "valor": iva
                        }
                    ],
                    "total_impuestos": iva,
                    "total_a_pagar": total,
                    "moneda": "COP",
                    "total_en_letras": f"{int(total)} PESOS M/CTE"
                },
                "observaciones": factura.get('observaciones') or 'Sin observaciones',
                "informacion_adicional": {
                    "nota": "Gracias por su compra",
                    "terminos_condiciones": "Esta factura es un título valor según Art. 772 del Código de Comercio",
                    "resolucion_dian": "000042 del 05 de mayo de 2020",
                    "rango_autorizacion": "FV-1 hasta FV-99999",
                    "vigencia_autorizacion": "2025-12-31"
                },
                "firma_electronica": {
                    "cufe": "A1B2C3D4E5F6G7H8I9J0K1L2M3N4O5P6Q7R8S9T0U1V2W3X4Y5Z6",
                    "algoritmo": "SHA-384",
                    "fecha_firma": datetime.now().isoformat(),
                    "politica_firma": "https://facturaelectronica.dian.gov.co/politicadefirma/v1"
                }
            }
            
            # Convertir a JSON con formato
            json_str = json.dumps(factura_json, indent=2, ensure_ascii=False)
            json_bytes = json_str.encode('utf-8')
            
            numero_factura = factura.get('numero_factura', 'N/A').replace('/', '-')
            fecha = factura.get('fecha_emision', 'N/A')[:10].replace('-', '')
            nombre_archivo = f"FV_{numero_factura}_{fecha}.json"
            
            st.download_button(
                label="📥 Descargar JSON",
                data=json_bytes,
                file_name=nombre_archivo,
                mime="application/json",
                type="primary",
                use_container_width=True
            )
            
            st.success("✅ Factura JSON generada exitosamente (Formato DIAN)")
            
            # Mostrar preview
            with st.expander("👁️ Vista previa del JSON (Compatibilidad DIAN)"):
                st.json(factura_json)
            
    except Exception as e:
        st.error(f"❌ Error al generar JSON: {str(e)}")
        st.exception(e)

def marcar_como_pagada(backend_url: str, id_factura: int):
    """Marcar factura como pagada"""
    
    try:
        with st.spinner("Actualizando estado de factura..."):
            response = requests.patch(
                f"{backend_url}/api/facturas/{id_factura}/pagar"
            )
        
        if response.status_code == 200:
            st.success("✅ Factura marcada como pagada")
            st.rerun()
        else:
            st.error(f"Error al actualizar factura: {response.status_code}")
            
    except Exception as e:
        st.error(f"Error al marcar como pagada: {e}")

def anular_factura(backend_url: str, id_factura: int):
    """Anular factura"""
    
    try:
        with st.spinner("Anulando factura..."):
            response = requests.patch(
                f"{backend_url}/api/facturas/{id_factura}/anular"
            )
        
        if response.status_code == 200:
            st.success("✅ Factura anulada")
            st.rerun()
        else:
            st.error(f"Error al anular factura: {response.status_code}")
            
    except Exception as e:
        st.error(f"Error al anular factura: {e}")

def reportes_ventas(backend_url: str):
    """Reportes y análisis de ventas"""
    
    st.subheader("📊 Reportes de Ventas")
    
    # Selector de tipo de reporte
    tipo_reporte = st.selectbox(
        "Tipo de reporte:",
        [
            "Ventas por período",
            "Ventas por cliente",
            "Ventas por producto",
            "Análisis de tendencias",
            "Estado de cartera"
        ]
    )
    
    # Filtros de fecha
    col1, col2 = st.columns(2)
    
    with col1:
        fecha_desde = st.date_input("Desde:", value=datetime.now().replace(day=1).date())
    
    with col2:
        fecha_hasta = st.date_input("Hasta:", value=datetime.now().date())
    
    if st.button("📊 Generar Reporte", width="stretch"):
        generar_reporte_ventas(backend_url, tipo_reporte, fecha_desde, fecha_hasta)

def generar_reporte_ventas(
    backend_url: str,
    tipo_reporte: str,
    fecha_desde: date,
    fecha_hasta: date
):
    """Generar reporte específico de ventas"""
    
    try:
        params = {
            "fecha_desde": fecha_desde.isoformat(),
            "fecha_hasta": fecha_hasta.isoformat()
        }
        
        with st.spinner(f"Generando {tipo_reporte}..."):
            if tipo_reporte == "Ventas por período":
                response = requests.get(f"{backend_url}/api/facturacion/reportes/ventas", params=params)
            elif tipo_reporte == "Ventas por cliente":
                response = requests.get(f"{backend_url}/api/facturacion/reportes/ventas-cliente", params=params)
            elif tipo_reporte == "Ventas por producto":
                response = requests.get(f"{backend_url}/api/facturacion/reportes/ventas-producto", params=params)
            elif tipo_reporte == "Análisis de tendencias":
                response = requests.get(f"{backend_url}/api/facturacion/reportes/tendencias-ventas", params=params)
            elif tipo_reporte == "Estado de cartera":
                response = requests.get(f"{backend_url}/api/facturacion/reportes/cuentas-por-cobrar", params=params)
        
        if response.status_code == 200:
            datos_reporte = response.json()
            mostrar_reporte_ventas(datos_reporte, tipo_reporte)
        else:
            st.error(f"Error al generar reporte: {response.status_code}")
            
    except Exception as e:
        st.error(f"Error al generar reporte: {e}")

def mostrar_reporte_ventas(datos: Dict[str, Any], tipo_reporte: str):
    """Mostrar reporte de ventas según el tipo"""
    
    st.markdown(f"### 📊 {tipo_reporte}")
    
    if tipo_reporte == "Ventas por período":
        mostrar_reporte_periodo(datos)
    elif tipo_reporte == "Ventas por cliente":
        mostrar_reporte_por_cliente(datos)
    elif tipo_reporte == "Ventas por producto":
        mostrar_reporte_por_producto(datos)
    elif tipo_reporte == "Análisis de tendencias":
        mostrar_reporte_tendencias(datos)
    elif tipo_reporte == "Estado de cartera":
        mostrar_reporte_cartera(datos)

def mostrar_reporte_periodo(datos: Dict[str, Any]):
    """Mostrar reporte de ventas por período"""
    
    # Información del período
    periodo = datos.get('periodo', {})
    st.info(f"📅 Período: {periodo.get('fecha_desde')} a {periodo.get('fecha_hasta')}")
    
    # Resumen principal
    resumen = datos.get('resumen', {})
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            "Facturas Emitidas",
            f"{resumen.get('cantidad_facturas', 0):,}"
        )
    
    with col2:
        st.metric(
            "Total Ventas",
            f"${resumen.get('total_ventas', 0):,.2f}"
        )
    
    with col3:
        st.metric(
            "Subtotal",
            f"${resumen.get('subtotal', 0):,.2f}"
        )
    
    with col4:
        st.metric(
            "IVA",
            f"${resumen.get('impuestos', 0):,.2f}"
        )
    
    st.divider()
    
    # Ventas por cliente
    st.markdown("#### 👥 Ventas por Cliente")
    ventas_cliente = datos.get('ventas_por_cliente', {})
    
    if ventas_cliente:
        clientes_df = []
        for cliente, info in ventas_cliente.items():
            clientes_df.append({
                'Cliente': cliente,
                'Facturas': info.get('cantidad_facturas', 0),
                'Subtotal': info.get('subtotal', 0),
                'IVA': info.get('impuestos', 0),
                'Total': info.get('total_ventas', 0)
            })
        
        import pandas as pd
        df = pd.DataFrame(clientes_df)
        df = df.sort_values('Total', ascending=False)
        
        st.dataframe(
            df.style.format({
                'Subtotal': '${:,.2f}',
                'IVA': '${:,.2f}',
                'Total': '${:,.2f}'
            }),
            width="stretch",
            hide_index=True
        )
    else:
        st.info("No hay ventas registradas en este período")
    
    st.divider()
    
    # Productos más vendidos
    st.markdown("#### 📦 Top 10 Productos Más Vendidos")
    productos_top = datos.get('productos_mas_vendidos', [])
    
    if productos_top:
        productos_df = []
        for idx, prod in enumerate(productos_top, 1):
            productos_df.append({
                '#': idx,
                'Producto': prod.get('producto', ''),
                'Cantidad': prod.get('cantidad_vendida', 0),
                'Total Vendido': prod.get('total_vendido', 0)
            })
        
        df_productos = pd.DataFrame(productos_df)
        
        st.dataframe(
            df_productos.style.format({
                'Cantidad': '{:,.0f}',
                'Total Vendido': '${:,.2f}'
            }),
            width="stretch",
            hide_index=True
        )
    else:
        st.info("No hay productos vendidos en este período")

def mostrar_reporte_por_cliente(datos: Dict[str, Any]):
    """Mostrar reporte de ventas agrupado por cliente"""
    
    import pandas as pd
    
    # Información del período
    periodo = datos.get('periodo', {})
    st.info(f"📅 Período: {periodo.get('fecha_desde')} a {periodo.get('fecha_hasta')}")
    
    # Resumen
    resumen = datos.get('resumen', {})
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Total Clientes", f"{resumen.get('total_clientes', 0):,}")
    
    with col2:
        st.metric("Total Ventas", f"${resumen.get('total_ventas', 0):,.2f}")
    
    with col3:
        st.metric("Promedio/Cliente", f"${resumen.get('promedio_por_cliente', 0):,.2f}")
    
    st.divider()
    
    # Detalle por cliente
    st.markdown("#### 👥 Detalle de Ventas por Cliente")
    
    clientes = datos.get('detalle_clientes', [])
    
    if clientes:
        clientes_df = []
        for cliente in clientes:
            clientes_df.append({
                'Código': cliente.get('codigo_cliente', ''),
                'NIT': cliente.get('nit', ''),
                'Cliente': cliente.get('nombre', ''),
                'Facturas': cliente.get('cantidad_facturas', 0),
                'Subtotal': cliente.get('subtotal', 0),
                'IVA': cliente.get('impuestos', 0),
                'Total': cliente.get('total_ventas', 0)
            })
        
        df = pd.DataFrame(clientes_df)
        
        st.dataframe(
            df.style.format({
                'Subtotal': '${:,.2f}',
                'IVA': '${:,.2f}',
                'Total': '${:,.2f}'
            }),
            width="stretch",
            hide_index=True
        )
        
        # Gráfico de barras top 10
        st.markdown("#### 📊 Top 10 Clientes")
        top_10 = df.head(10)
        
        import plotly.express as px
        fig = px.bar(
            top_10,
            x='Cliente',
            y='Total',
            title='Top 10 Clientes por Ventas',
            labels={'Total': 'Ventas Totales ($)', 'Cliente': 'Cliente'}
        )
        st.plotly_chart(fig, use_container_width=True)
        
    else:
        st.info("No hay ventas registradas en este período")

def mostrar_reporte_por_producto(datos: Dict[str, Any]):
    """Mostrar reporte de ventas agrupado por producto"""
    
    import pandas as pd
    
    # Información del período
    periodo = datos.get('periodo', {})
    st.info(f"📅 Período: {periodo.get('fecha_desde')} a {periodo.get('fecha_hasta')}")
    
    # Resumen
    resumen = datos.get('resumen', {})
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Productos", f"{resumen.get('total_productos', 0):,}")
    
    with col2:
        st.metric("Cantidad Vendida", f"{resumen.get('cantidad_total_vendida', 0):,.0f}")
    
    with col3:
        st.metric("Total Ventas", f"${resumen.get('total_ventas', 0):,.2f}")
    
    with col4:
        st.metric("Ticket Promedio", f"${resumen.get('ticket_promedio', 0):,.2f}")
    
    st.divider()
    
    # Detalle por producto
    st.markdown("#### 📦 Detalle de Ventas por Producto")
    
    productos = datos.get('detalle_productos', [])
    
    if productos:
        productos_df = []
        for prod in productos:
            productos_df.append({
                'Código': prod.get('codigo_producto', ''),
                'Producto': prod.get('nombre', ''),
                'Precio Venta': prod.get('precio_venta', 0),
                'Cantidad': prod.get('cantidad_vendida', 0),
                'Subtotal': prod.get('subtotal', 0),
                'IVA': prod.get('impuestos', 0),
                'Total': prod.get('total_vendido', 0),
                'Precio Prom.': prod.get('precio_promedio', 0)
            })
        
        df = pd.DataFrame(productos_df)
        
        st.dataframe(
            df.style.format({
                'Precio Venta': '${:,.2f}',
                'Cantidad': '{:,.0f}',
                'Subtotal': '${:,.2f}',
                'IVA': '${:,.2f}',
                'Total': '${:,.2f}',
                'Precio Prom.': '${:,.2f}'
            }),
            width="stretch",
            hide_index=True
        )
        
        # Gráfico de productos más vendidos
        st.markdown("#### 📊 Top 10 Productos Más Vendidos")
        top_10 = df.head(10)
        
        import plotly.express as px
        fig = px.bar(
            top_10,
            x='Producto',
            y='Cantidad',
            title='Top 10 Productos por Cantidad',
            labels={'Cantidad': 'Unidades Vendidas', 'Producto': 'Producto'}
        )
        st.plotly_chart(fig, use_container_width=True)
        
    else:
        st.info("No hay productos vendidos en este período")

def mostrar_reporte_tendencias(datos: Dict[str, Any]):
    """Mostrar reporte de tendencias de ventas"""
    
    import pandas as pd
    
    # Información del período
    periodo = datos.get('periodo', {})
    st.info(f"📅 Período: {periodo.get('fecha_desde')} a {periodo.get('fecha_hasta')}")
    
    # Resumen
    resumen = datos.get('resumen', {})
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Ventas", f"${resumen.get('total_ventas', 0):,.2f}")
    
    with col2:
        st.metric("Total Facturas", f"{resumen.get('total_facturas', 0):,}")
    
    with col3:
        st.metric("Promedio Diario", f"${resumen.get('promedio_diario', 0):,.2f}")
    
    with col4:
        st.metric("Promedio/Factura", f"${resumen.get('promedio_por_factura', 0):,.2f}")
    
    st.divider()
    
    # Mejor y peor día
    col1, col2 = st.columns(2)
    
    mejor_dia = datos.get('mejor_dia', {})
    peor_dia = datos.get('peor_dia', {})
    
    with col1:
        if mejor_dia:
            st.success("🏆 Mejor Día de Ventas")
            st.markdown(f"**Fecha:** {mejor_dia.get('fecha', '')}")
            st.markdown(f"**Facturas:** {mejor_dia.get('cantidad_facturas', 0)}")
            st.markdown(f"**Total:** ${mejor_dia.get('total', 0):,.2f}")
    
    with col2:
        if peor_dia:
            st.warning("📉 Día con Menores Ventas")
            st.markdown(f"**Fecha:** {peor_dia.get('fecha', '')}")
            st.markdown(f"**Facturas:** {peor_dia.get('cantidad_facturas', 0)}")
            st.markdown(f"**Total:** ${peor_dia.get('total', 0):,.2f}")
    
    st.divider()
    
    # Gráfico de tendencias
    st.markdown("#### 📈 Evolución de Ventas Diarias")
    
    ventas_diarias = datos.get('ventas_diarias', [])
    
    if ventas_diarias:
        df = pd.DataFrame(ventas_diarias)
        
        import plotly.graph_objects as go
        
        fig = go.Figure()
        
        fig.add_trace(go.Scatter(
            x=df['fecha'],
            y=df['total'],
            mode='lines+markers',
            name='Total',
            line=dict(color='#1f77b4', width=2)
        ))
        
        fig.update_layout(
            title='Ventas Diarias',
            xaxis_title='Fecha',
            yaxis_title='Total ($)',
            hovermode='x unified'
        )
        
        st.plotly_chart(fig, use_container_width=True)
        
        # Tabla de datos
        st.markdown("#### 📋 Datos Diarios")
        st.dataframe(
            df.style.format({
                'subtotal': '${:,.2f}',
                'impuestos': '${:,.2f}',
                'total': '${:,.2f}'
            }),
            width="stretch",
            hide_index=True
        )
    else:
        st.info("No hay ventas registradas en este período")

def mostrar_reporte_cartera(datos: Dict[str, Any]):
    """Mostrar reporte de cuentas por cobrar"""
    
    import pandas as pd
    
    # Fecha de corte
    fecha_corte = datos.get('fecha_corte', '')
    st.info(f"📅 Fecha de Corte: {fecha_corte}")
    
    # Resumen
    col1, col2 = st.columns(2)
    
    with col1:
        st.metric(
            "Total Cuentas por Cobrar",
            f"${datos.get('total_cuentas_por_cobrar', 0):,.2f}"
        )
    
    with col2:
        st.metric(
            "Facturas Pendientes",
            f"{datos.get('cantidad_facturas_pendientes', 0):,}"
        )
    
    st.divider()
    
    # Resumen por estado
    st.markdown("#### 📊 Resumen por Estado de Cobranza")
    
    resumen_estado = datos.get('resumen_por_estado', {})
    
    if resumen_estado:
        estados_df = []
        estados_labels = {
            'AL_DIA': '✅ Al Día',
            'VENCIDA_30': '⚠️ Vencida 1-30 días',
            'VENCIDA_60': '🟡 Vencida 31-60 días',
            'VENCIDA_90': '🟠 Vencida 61-90 días',
            'VENCIDA_MAS_90': '🔴 Vencida +90 días'
        }
        
        for estado, info in resumen_estado.items():
            if info['cantidad'] > 0:
                estados_df.append({
                    'Estado': estados_labels.get(estado, estado),
                    'Cantidad': info['cantidad'],
                    'Total': info['total']
                })
        
        if estados_df:
            df_estados = pd.DataFrame(estados_df)
            
            col1, col2 = st.columns([2, 1])
            
            with col1:
                st.dataframe(
                    df_estados.style.format({
                        'Total': '${:,.2f}'
                    }),
                    width="stretch",
                    hide_index=True
                )
            
            with col2:
                import plotly.express as px
                fig = px.pie(
                    df_estados,
                    values='Total',
                    names='Estado',
                    title='Distribución de Cartera'
                )
                st.plotly_chart(fig, use_container_width=True)
    
    st.divider()
    
    # Detalle de facturas pendientes
    st.markdown("#### 📋 Detalle de Facturas Pendientes")
    
    detalle = datos.get('detalle', [])
    
    if detalle:
        facturas_df = []
        for factura in detalle:
            cliente = factura.get('cliente', {})
            
            estado_labels = {
                'AL_DIA': '✅ Al Día',
                'VENCIDA_30': '⚠️ 1-30 días',
                'VENCIDA_60': '🟡 31-60 días',
                'VENCIDA_90': '🟠 61-90 días',
                'VENCIDA_MAS_90': '🔴 +90 días'
            }
            
            facturas_df.append({
                'Factura': factura.get('numero_factura', ''),
                'Cliente': cliente.get('nombre', ''),
                'NIT': cliente.get('nit', ''),
                'Emisión': factura.get('fecha_factura', ''),
                'Vencimiento': factura.get('fecha_vencimiento', ''),
                'Días Venc.': factura.get('dias_vencimiento', 0),
                'Total': factura.get('total', 0),
                'Estado': estado_labels.get(factura.get('estado_cobranza', ''), '')
            })
        
        df = pd.DataFrame(facturas_df)
        
        st.dataframe(
            df.style.format({
                'Total': '${:,.2f}'
            }),
            width="stretch",
            hide_index=True
        )
    else:
        st.success("✅ No hay facturas pendientes por cobrar")

def configuracion_facturacion(backend_url: str):
    """Configuración del módulo de facturación"""
    
    st.subheader("⚙️ Configuración de Facturación")
    
    # Tabs para diferentes configuraciones
    tab_clientes, tab_productos, tab_parametros = st.tabs([
        "👥 Gestión de Clientes",
        "📦 Gestión de Productos", 
        "⚙️ Parámetros"
    ])
    
    with tab_clientes:
        gestion_clientes(backend_url)
    
    with tab_productos:
        gestion_productos(backend_url)
    
    with tab_parametros:
        configurar_parametros(backend_url)

def gestion_clientes(backend_url: str):
    """Gestión de clientes"""
    
    st.markdown("#### 👥 Gestión de Clientes")
    
    # Botones de acción
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("➕ Agregar Cliente", width="stretch"):
            st.session_state.mostrar_form_cliente = True
    
    with col2:
        if st.button("🔄 Actualizar Lista", width="stretch"):
            st.rerun()
    
    # Formulario para nuevo cliente
    if st.session_state.get('mostrar_form_cliente', False):
        with st.form("form_nuevo_cliente"):
            st.markdown("**➕ Nuevo Cliente**")
            
            col1, col2 = st.columns(2)
            
            with col1:
                codigo_cliente = st.text_input("Código Cliente*:", help="Código único del cliente")
                nombre = st.text_input("Nombre/Razón Social*:")
                nit = st.text_input("NIT/CC*:")
                email = st.text_input("Email:")
            
            with col2:
                telefono = st.text_input("Teléfono:")
                direccion = st.text_area("Dirección:")
                tipo_cliente = st.selectbox("Tipo:", ["Empresa", "Persona Natural"])
                estado = st.selectbox("Estado:", ["ACTIVO", "INACTIVO", "BLOQUEADO"], index=0)
            
            col1, col2 = st.columns(2)
            
            with col1:
                if st.form_submit_button("💾 Guardar Cliente", width="stretch"):
                    if codigo_cliente and nombre and nit:
                        crear_cliente(
                            backend_url, codigo_cliente, nombre, nit, 
                            email, telefono, direccion, tipo_cliente, estado
                        )
                    else:
                        st.error("❌ Complete los campos obligatorios")
            
            with col2:
                if st.form_submit_button("❌ Cancelar"):
                    st.session_state.mostrar_form_cliente = False
                    st.rerun()
    
    # Lista de clientes existentes
    mostrar_lista_clientes(backend_url)

def crear_cliente(
    backend_url: str, codigo: str, nombre: str, nit: str,
    email: str, telefono: str, direccion: str, tipo: str, estado: str
):
    """Crear nuevo cliente"""
    
    try:
        # Mapear el tipo de cliente al formato del backend
        tipo_cliente_backend = "PERSONA_JURIDICA" if tipo == "Empresa" else "PERSONA_NATURAL"
        estado_backend = estado  # Ya viene en formato correcto: ACTIVO, INACTIVO, BLOQUEADO
        
        datos_cliente = {
            "codigo_cliente": codigo if codigo else None,
            "nombre": nombre,
            "nit": nit if nit else None,
            "email": email if email else None,
            "telefono_principal": telefono if telefono else None,
            "direccion": direccion if direccion else None,
            "tipo_cliente": tipo_cliente_backend,
            "estado_cliente": estado_backend,
            "usuario_creacion": "SISTEMA"  # Campo requerido
        }
        
        with st.spinner("Creando cliente..."):
            response = requests.post(f"{backend_url}/api/facturacion/clientes", json=datos_cliente)
        
        if response.status_code in [200, 201]:
            st.success("✅ Cliente creado exitosamente")
            st.session_state.mostrar_form_cliente = False
            st.rerun()
        else:
            error_detail = response.json().get('detail', 'Error desconocido')
            st.error(f"❌ Error al crear cliente: {error_detail}")
            
    except Exception as e:
        st.error(f"❌ Error al crear cliente: {e}")

def mostrar_lista_clientes(backend_url: str):
    """Mostrar lista de clientes"""
    
    try:
        with st.spinner("Cargando clientes..."):
            response = requests.get(f"{backend_url}/api/facturacion/clientes")
        
        if response.status_code == 200:
            clientes = response.json()
            
            if clientes:
                st.markdown("**📋 Lista de Clientes**")
                
                df_clientes = pd.DataFrame(clientes)
                
                # Formatear para mostrar - usar los campos correctos del schema
                columnas_disponibles = []
                for col in ['codigo_cliente', 'nombre', 'nit', 'tipo_cliente', 'estado_cliente']:
                    if col in df_clientes.columns:
                        columnas_disponibles.append(col)
                
                if columnas_disponibles:
                    df_display = df_clientes[columnas_disponibles].copy()
                    # Renombrar columnas para mejor visualización
                    nombres_columnas = {
                        'codigo_cliente': 'Código',
                        'nombre': 'Nombre',
                        'nit': 'NIT',
                        'tipo_cliente': 'Tipo',
                        'estado_cliente': 'Estado'
                    }
                    df_display.columns = [nombres_columnas.get(col, col) for col in columnas_disponibles]
                    
                    st.dataframe(df_display, width="stretch", hide_index=True)
                else:
                    st.info("No hay datos suficientes para mostrar")
            else:
                st.info("📭 No hay clientes registrados")
        else:
            st.error(f"Error al cargar clientes: {response.status_code}")
            
    except Exception as e:
        st.error(f"Error al cargar clientes: {e}")

def gestion_productos(backend_url: str):
    """Gestión de productos"""
    
    st.markdown("#### 📦 Gestión de Productos")
    
    # Botones de acción
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("➕ Agregar Producto", width="stretch"):
            st.session_state.mostrar_form_producto = True
    
    with col2:
        if st.button("🔄 Actualizar Lista", width="stretch", key="refresh_productos"):
            st.rerun()
    
    # Formulario para nuevo producto
    if st.session_state.get('mostrar_form_producto', False):
        with st.form("form_nuevo_producto"):
            st.markdown("**➕ Nuevo Producto/Servicio**")
            
            col1, col2 = st.columns(2)
            
            with col1:
                codigo_producto = st.text_input("Código Producto*:")
                nombre = st.text_input("Nombre*:")
                precio_unitario = st.number_input("Precio Unitario*:", min_value=0.0, step=0.01)
                tipo_producto = st.selectbox("Tipo:", ["PRODUCTO", "SERVICIO", "COMBO"])
            
            with col2:
                descripcion = st.text_area("Descripción:")
                categoria = st.text_input("Categoría:")
                estado = st.selectbox("Estado:", ["ACTIVO", "INACTIVO", "DESCONTINUADO"], index=0)
            
            col1, col2 = st.columns(2)
            
            with col1:
                if st.form_submit_button("💾 Guardar Producto", width="stretch"):
                    if codigo_producto and nombre and precio_unitario > 0:
                        crear_producto(
                            backend_url, codigo_producto, nombre, precio_unitario,
                            descripcion, categoria, tipo_producto, estado
                        )
                    else:
                        st.error("❌ Complete los campos obligatorios")
            
            with col2:
                if st.form_submit_button("❌ Cancelar", key="cancel_producto"):
                    st.session_state.mostrar_form_producto = False
                    st.rerun()
    
    # Lista de productos existentes
    mostrar_lista_productos(backend_url)

def crear_producto(
    backend_url: str, codigo: str, nombre: str, precio_unitario: float,
    descripcion: str, categoria: str, tipo: str, estado: str
):
    """Crear nuevo producto"""
    
    try:
        datos_producto = {
            "codigo_producto": codigo,
            "nombre": nombre,
            "descripcion": descripcion if descripcion else None,
            "tipo_producto": tipo,  # Ya viene en formato correcto (PRODUCTO, SERVICIO, COMBO)
            "precio_venta": float(precio_unitario),
            "precio_compra": 0.0,  # Valor por defecto
            "aplica_iva": True,  # Valor por defecto
            "porcentaje_iva": 13.0,  # Valor por defecto para El Salvador
            "estado_producto": estado,  # Ya viene en formato correcto (ACTIVO, INACTIVO, DESCONTINUADO)
            "categoria_producto": categoria if categoria else None
        }
        
        with st.spinner("Creando producto..."):
            response = requests.post(f"{backend_url}/api/facturacion/productos", json=datos_producto)
        
        if response.status_code in [200, 201]:
            st.success("✅ Producto creado exitosamente")
            st.session_state.mostrar_form_producto = False
            st.rerun()
        else:
            error_detail = response.json().get('detail', 'Error desconocido')
            st.error(f"❌ Error al crear producto: {error_detail}")
            
    except Exception as e:
        st.error(f"❌ Error al crear producto: {e}")

def mostrar_lista_productos(backend_url: str):
    """Mostrar lista de productos"""
    
    try:
        with st.spinner("Cargando productos..."):
            response = requests.get(f"{backend_url}/api/facturacion/productos")
        
        if response.status_code == 200:
            productos = response.json()
            
            if productos:
                st.markdown("**📦 Lista de Productos/Servicios**")
                
                df_productos = pd.DataFrame(productos)
                
                # Formatear para mostrar - usar los campos correctos del schema
                columnas_disponibles = []
                for col in ['codigo_producto', 'nombre', 'precio_venta', 'tipo_producto', 'estado_producto']:
                    if col in df_productos.columns:
                        columnas_disponibles.append(col)
                
                if columnas_disponibles:
                    df_display = df_productos[columnas_disponibles].copy()
                    
                    # Formatear precio si existe
                    if 'precio_venta' in df_display.columns:
                        df_display['precio_venta'] = df_display['precio_venta'].apply(lambda x: f"${float(x):,.2f}")
                    
                    # Renombrar columnas para mejor visualización
                    nombres_columnas = {
                        'codigo_producto': 'Código',
                        'nombre': 'Nombre',
                        'precio_venta': 'Precio',
                        'tipo_producto': 'Tipo',
                        'estado_producto': 'Estado'
                    }
                    df_display.columns = [nombres_columnas.get(col, col) for col in columnas_disponibles]
                    
                    st.dataframe(df_display, width="stretch", hide_index=True)
                else:
                    st.info("No hay datos suficientes para mostrar")
            else:
                st.info("📭 No hay productos registrados")
        else:
            st.error(f"Error al cargar productos: {response.status_code}")
            
    except Exception as e:
        st.error(f"Error al cargar productos: {e}")

def configurar_parametros(backend_url: str):
    """Configurar parámetros de facturación"""
    
    st.markdown("#### ⚙️ Parámetros de Facturación")
    
    # Cargar configuración existente
    try:
        with st.spinner("Cargando configuración..."):
            response = requests.get(f"{backend_url}/api/facturacion/configuracion")
        
        if response.status_code == 200:
            config = response.json()
        else:
            config = None
            st.warning("⚠️ No hay configuración guardada. Se usarán valores por defecto.")
    except Exception as e:
        config = None
        st.error(f"Error al cargar configuración: {e}")
    
    # Valores por defecto o cargados
    empresa_nit_default = config.get('empresa_nit', '000000000-0') if config else '000000000-0'
    empresa_nombre_default = config.get('empresa_nombre', 'Mi Empresa') if config else 'Mi Empresa'
    empresa_direccion_default = config.get('empresa_direccion', '') if config else ''
    empresa_telefono_default = config.get('empresa_telefono', '') if config else ''
    empresa_email_default = config.get('empresa_email', '') if config else ''
    empresa_web_default = config.get('empresa_web', '') if config else ''
    
    iva_default = float(config.get('iva_porcentaje', 13.0)) if config else 13.0
    retefuente_default = float(config.get('retefuente_porcentaje', 0.0)) if config else 0.0
    reteica_default = float(config.get('reteica_porcentaje', 0.0)) if config else 0.0
    
    prefijo_default = config.get('prefijo_factura', 'FV') if config else 'FV'
    numero_inicial_default = config.get('numero_inicial', 1) if config else 1
    numero_actual_default = config.get('numero_actual', 1) if config else 1
    
    # Formulario de configuración
    with st.form("form_configuracion"):
        # Parámetros generales
        st.markdown("**🏢 Información de la Empresa**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            empresa_nit = st.text_input("NIT Empresa*:", value=empresa_nit_default)
            empresa_nombre = st.text_input("Razón Social*:", value=empresa_nombre_default)
            empresa_direccion = st.text_area("Dirección:", value=empresa_direccion_default)
        
        with col2:
            empresa_telefono = st.text_input("Teléfono:", value=empresa_telefono_default)
            empresa_email = st.text_input("Email:", value=empresa_email_default)
            empresa_web = st.text_input("Sitio Web:", value=empresa_web_default)
        
        st.markdown("**💰 Parámetros Fiscales**")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            iva_porcentaje = st.number_input("IVA (%):", value=iva_default, min_value=0.0, max_value=100.0, step=0.1)
        
        with col2:
            retefuente_porcentaje = st.number_input("Retención Fuente (%):", value=retefuente_default, min_value=0.0, max_value=100.0, step=0.1)
        
        with col3:
            reteica_porcentaje = st.number_input("ReteICA (%):", value=reteica_default, min_value=0.0, max_value=100.0, step=0.001, format="%.3f")
        
        # Numeración de facturas
        st.markdown("**🔢 Numeración de Facturas**")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            prefijo_factura = st.text_input("Prefijo:", value=prefijo_default)
        
        with col2:
            numero_inicial = st.number_input("Número Inicial:", value=numero_inicial_default, min_value=1, step=1)
        
        with col3:
            numero_actual = st.number_input("Número Actual:", value=numero_actual_default, min_value=1, step=1)
        
        st.info("ℹ️ El número actual se incrementa automáticamente con cada factura emitida.")
        
        # Botón de guardado
        if st.form_submit_button("💾 Guardar Configuración", width="stretch"):
            if empresa_nit and empresa_nombre:
                guardar_configuracion_facturacion(
                    backend_url,
                    empresa_nit, empresa_nombre, empresa_direccion,
                    empresa_telefono, empresa_email, empresa_web,
                    iva_porcentaje, retefuente_porcentaje, reteica_porcentaje,
                    prefijo_factura, numero_inicial, numero_actual,
                    config is not None
                )
            else:
                st.error("❌ Complete los campos obligatorios: NIT y Razón Social")

def guardar_configuracion_facturacion(
    backend_url: str,
    empresa_nit: str, empresa_nombre: str, empresa_direccion: str,
    empresa_telefono: str, empresa_email: str, empresa_web: str,
    iva_porcentaje: float, retefuente_porcentaje: float, reteica_porcentaje: float,
    prefijo_factura: str, numero_inicial: int, numero_actual: int,
    existe_config: bool
):
    """Guardar o actualizar configuración de facturación"""
    
    try:
        datos_config = {
            "empresa_nit": empresa_nit,
            "empresa_nombre": empresa_nombre,
            "empresa_direccion": empresa_direccion if empresa_direccion else None,
            "empresa_telefono": empresa_telefono if empresa_telefono else None,
            "empresa_email": empresa_email if empresa_email else None,
            "empresa_web": empresa_web if empresa_web else None,
            "iva_porcentaje": iva_porcentaje,
            "retefuente_porcentaje": retefuente_porcentaje,
            "reteica_porcentaje": reteica_porcentaje,
            "prefijo_factura": prefijo_factura,
            "numero_inicial": numero_inicial,
            "numero_actual": numero_actual,
            "usuario_actualizacion": "SISTEMA"
        }
        
        with st.spinner("Guardando configuración..."):
            if existe_config:
                # Actualizar configuración existente
                response = requests.put(
                    f"{backend_url}/api/facturacion/configuracion",
                    json=datos_config
                )
            else:
                # Crear nueva configuración
                response = requests.post(
                    f"{backend_url}/api/facturacion/configuracion",
                    json=datos_config
                )
        
        if response.status_code in [200, 201]:
            st.success("✅ Configuración guardada exitosamente")
            st.info("🔄 Los cambios se aplicarán en las próximas operaciones")
            st.rerun()
        else:
            error_detail = response.json().get('detail', 'Error desconocido')
            st.error(f"❌ Error al guardar configuración: {error_detail}")
            
    except Exception as e:
        st.error(f"❌ Error al guardar configuración: {e}")
